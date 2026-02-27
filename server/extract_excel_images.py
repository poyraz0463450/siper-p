"""
Excel Image Extractor
Extracts embedded images from an Excel file and saves them with names based on adjacent cell values
"""
import openpyxl
from openpyxl.drawing.image import Image
import os
import sys
from pathlib import Path

def extract_images_from_excel(excel_path, output_dir='extracted_images'):
    """
    Extract all images from an Excel file and match them to codes
    
    Args:
        excel_path: Path to the Excel file
        output_dir: Directory to save extracted images
    """
    # Create output directory if it doesn't exist
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Load workbook
    print(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    
    # Process first sheet
    sheet = wb.active
    print(f"Processing sheet: {sheet.title}")
    
    # Get all images from the sheet
    images = sheet._images if hasattr(sheet, '_images') else []
    
    if not images:
        print("No images found in the worksheet")
        # Try to show worksheet data for debugging
        print("\nFirst few rows of data:")
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
            print(f"Row {row_idx}: {row}")
        return
    
    print(f"Found {len(images)} images in the worksheet")
    
    extracted_count = 0
    
    # Extract each image
    for idx, img in enumerate(images, 1):
        try:
            # Get the image anchor position
            anchor = img.anchor
            
            # Try to determine the row where the image is located
            # This helps us find the corresponding operation code
            if hasattr(anchor, '_from'):
                row_idx = anchor._from.row + 1  # openpyxl uses 0-based indexing
                col_idx = anchor._from.col + 1
                
                # Try to find operation code in nearby cells
                # Common patterns: code might be in the same row, adjacent columns
                operation_code = None
                
                # Check current row, various columns
                for check_col in range(max(1, col_idx - 5), col_idx + 6):
                    try:
                        cell_value = sheet.cell(row=row_idx, column=check_col).value
                        if cell_value and isinstance(cell_value, str):
                            # Look for operation code pattern (contains hyphen and letters)
                            if '-' in cell_value and any(c.isalpha() for c in cell_value):
                                operation_code = cell_value.strip()
                                break
                    except:
                        continue
                
                # Generate filename
                if operation_code:
                    # Clean the operation code for use in filename
                    safe_code = "".join(c for c in operation_code if c.isalnum() or c in ('-', '_')).rstrip()
                    filename = f"{safe_code}.png"
                else:
                    filename = f"image_{idx}.png"
                
                # Save the image
                output_path = os.path.join(output_dir, filename)
                
                # Get image data
                img_data = img._data
                if img_data:
                    with open(output_path, 'wb') as f:
                        f.write(img_data())
                    
                    print(f"✓ Extracted: {filename} (Row {row_idx}, matched code: {operation_code or 'N/A'})")
                    extracted_count += 1
                        
        except Exception as e:
            print(f"✗ Error extracting image {idx}: {str(e)}")
            continue
    
    print(f"\nExtraction complete: {extracted_count} images saved to '{output_dir}'")
    
    # Also print a data preview for reference
    print("\nSample data from Excel (first 5 rows):")
    for row_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), 1):
        print(f"Row {row_idx}: {row[:10]}")  # First 10 columns

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_excel_images.py <path_to_excel_file> [output_directory]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_directory = sys.argv[2] if len(sys.argv) > 2 else 'extracted_images'
    
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)
    
    try:
        extract_images_from_excel(excel_file, output_directory)
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
